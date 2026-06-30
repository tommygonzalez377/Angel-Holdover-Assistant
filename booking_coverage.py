"""
Booking Coverage — committed (Mica) vs actual (showtimes) data layer.

Self-contained so it can be imported into the Angel Holdover Assistant
(launcher.py) later. All Snowflake creds come from environment variables:

  SNOWFLAKE_ACCOUNT   (default RIDEDVP-ANGEL)
  SNOWFLAKE_USER      (default tommy.gonzalez@angel.com)
  SNOWFLAKE_WAREHOUSE (default theatrical)
  SNOWFLAKE_ROLE      (default TOMMYG_RAW_AIRBYTE_READ)
  SNOWFLAKE_PAT       (the programmatic access token — required)
"""
import os
import re
import json
import threading
import urllib.request
import urllib.error
from datetime import datetime, timezone, date, timedelta
import snowflake.connector as sc

# Under-running thresholds for a "Clean" (full-slate) commitment over the window.
UNDER_DAYS = 3        # plays this many distinct days or fewer
UNDER_SHOWTIMES = 10  # OR runs this many showtimes or fewer

# Snowfall (Oracle) API base for venue tags — reachable over Twingate.
SNOWFALL_API = os.environ.get("SNOWFALL_API", "https://qa.angelstudios.com/api")

# Tags that explain a legitimate zero-online-showtimes state, so a "no_show"
# with one of these is reclassified to "expected" (not a real miss).
EXPLAIN_TAGS = {
    "no-online-sales", "no-showtimes", "no-website",
    "no-early-showtimes", "1-week-out-venue", "2-week-out-venue",
}

_lock = threading.Lock()
_conn = None
_bridge = None  # cached ORACLEV3 mica->provider_id maps (loaded once)

# Result cache so the page loads instantly between scheduled refreshes.
CACHE_FILE = os.path.join(os.path.dirname(__file__), "coverage_cache.json")
DEFAULT_WINDOW_DAYS = 6
_cache_lock = threading.Lock()
_cache = None


def _norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _connect():
    """Connect via key-pair (SNOWFLAKE_PRIVATE_KEY_B64 / _PATH) if present, else a
    PAT in SNOWFLAKE_PAT. Honors SNOWFLAKE_HOST for the load-balanced endpoint."""
    base = dict(
        account=os.environ.get("SNOWFLAKE_ACCOUNT", "RIDEDVP-ANGEL"),
        user=os.environ.get("SNOWFLAKE_USER", "tommy.gonzalez@angel.com"),
        warehouse=os.environ.get("SNOWFLAKE_WAREHOUSE", "theatrical"),
        role=os.environ.get("SNOWFLAKE_ROLE", "TOMMYG_RAW_AIRBYTE_READ"),
        login_timeout=30,
    )
    host = os.environ.get("SNOWFLAKE_HOST")  # e.g. ridedvp-angel.yukicomputing.com (load-balanced)
    if host:
        base["host"] = host
    key_b64 = os.environ.get("SNOWFLAKE_PRIVATE_KEY_B64")
    key_path = os.environ.get("SNOWFLAKE_PRIVATE_KEY_PATH")
    if key_b64 or key_path:
        import base64 as _b64
        from cryptography.hazmat.primitives import serialization
        pem = _b64.b64decode(key_b64) if key_b64 else open(key_path, "rb").read()
        passphrase = os.environ.get("SNOWFLAKE_PRIVATE_KEY_PASSPHRASE")
        pkey = serialization.load_pem_private_key(
            pem, password=passphrase.encode() if passphrase else None)
        der = pkey.private_bytes(serialization.Encoding.DER,
                                 serialization.PrivateFormat.PKCS8,
                                 serialization.NoEncryption())
        return sc.connect(private_key=der, **base)
    return sc.connect(password=os.environ["SNOWFLAKE_PAT"], **base)


def _cursor():
    """One shared connection; reconnect if it dropped."""
    global _conn
    if _conn is None:
        _conn = _connect()
    try:
        cur = _conn.cursor()
        cur.execute("SELECT 1")
        cur.fetchall()
        return _conn.cursor()
    except Exception:
        _conn = _connect()
        return _conn.cursor()


def _load_bridge(cur):
    """ORACLEV3_VENUES maps: normalized mica_name -> provider_id, atom_id -> provider_id."""
    global _bridge
    if _bridge is not None:
        return _bridge
    cur.execute("""
        SELECT mica_name, NULLIF(atom_tickets_id,''), provider_id
        FROM RAW_AIRBYTE.ORACLE_SERVICE.ORACLEV3_VENUES
    """)
    ov_name, ov_atom = {}, {}
    for mn, at, pid in cur.fetchall():
        if pid is None:
            continue
        pid = str(pid)
        if mn:
            ov_name.setdefault(_norm(mn), pid)
        if at:
            ov_atom.setdefault(str(at), pid)
    _bridge = (ov_name, ov_atom)
    return _bridge


def get_titles():
    """Upcoming titles that have a current Mica plan, with opening date."""
    with _lock:
        cur = _cursor()
        cur.execute("""
            SELECT m.production AS slug, p.name AS name,
                   TO_CHAR(p.theatrical_start_date,'YYYY-MM-DD') AS start_date
            FROM (SELECT DISTINCT production FROM RAW_AIRBYTE.ORACLE_SERVICE.IMPORTED_MICA_REPORT
                  WHERE import_date > DATEADD(day,-30,CURRENT_DATE())) m
            JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.PRODUCTIONS p ON p.theatrical_slug = m.production
            QUALIFY ROW_NUMBER() OVER (PARTITION BY m.production ORDER BY p.theatrical_start_date DESC NULLS LAST)=1
            ORDER BY start_date NULLS LAST
        """)
        return [{"slug": s, "name": n, "start_date": d} for s, n, d in cur.fetchall()]


def fetch_tags(slug, venue_uuids):
    """Venue tags from the snowfall API, keyed by monolith venue UUID.
    Returns {uuid: [tag, ...]}. Resilient: returns {} if the API is unreachable."""
    uuids = [u for u in venue_uuids if u]
    if not uuids:
        return {}
    out = {}
    url = f"{SNOWFALL_API}/tags/bulk-groupby/{slug}"
    for i in range(0, len(uuids), 2000):          # endpoint caps at 6000; stay well under
        batch = uuids[i:i + 2000]
        try:
            req = urllib.request.Request(
                url, data=json.dumps({"venueIds": batch}).encode(),
                headers={"Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=25) as resp:
                data = json.loads(resp.read().decode())
                out.update(data.get("tags", {}))
        except Exception as e:
            print(f"[tags] fetch failed ({e}); continuing without tags", flush=True)
            break
    return out


def fetch_coverage(slug, start, end):
    """Return committed-vs-actual coverage for one title + date window.

    Returns dict: {title, window, kpis, by_booker, venues}
    venues = list of per-Agreed-venue rows with a flag.
    """
    with _lock:
        cur = _cursor()
        ov_name, ov_atom = _load_bridge(cur)

        # committed (Agreed) venues from latest Mica import
        cur.execute(f"""
            SELECT venue_name, NULLIF(atom_tickets_id,'') AS atom_id,
                   screenings AS playtype, screens, status, bookers, buyer, buyer_email, city, state
            FROM RAW_AIRBYTE.ORACLE_SERVICE.IMPORTED_MICA_REPORT m
            WHERE production=%(slug)s
              AND DATE(import_date)=(SELECT MAX(DATE(import_date))
                                     FROM RAW_AIRBYTE.ORACLE_SERVICE.IMPORTED_MICA_REPORT
                                     WHERE production=%(slug)s)
              AND status IN ('Agreed','Booked')
            QUALIFY ROW_NUMBER() OVER (PARTITION BY venue_name ORDER BY import_date DESC)=1
        """, {"slug": slug})
        mica = [dict(venue=r[0], atom=r[1], playtype=r[2], screens=r[3],
                     status=r[4], booker=r[5], buyer=r[6], buyer_email=r[7],
                     city=r[8], state=r[9])
                for r in cur.fetchall()]

        # actual showtimes aggregated by provider_venue_id, with per-day daypart
        # coverage. Dayparts (by local start hour): matinee 7:00am-3:59pm,
        # prime 4:00pm-9:59pm, late 10:00pm+. A day is "clean" when it has both a
        # matinee AND a prime showstart. Aggregate by pvid (not name) to avoid the
        # stale-duplicate-venue undercount.
        cur.execute("""
            WITH daily AS (
              SELECT CAST(v.provider_venue_id AS STRING) AS pvid,
                     DATE(s.local_start_time) AS d,
                     COUNT_IF(HOUR(s.local_start_time) BETWEEN 7 AND 15)  AS mat,
                     COUNT_IF(HOUR(s.local_start_time) BETWEEN 16 AND 21) AS prime,
                     COUNT_IF(HOUR(s.local_start_time) >= 22 OR HOUR(s.local_start_time) < 7) AS late
              FROM RAW_AIRBYTE.THEATRICAL_MONOLITH.SHOWTIMES s
              JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.VENUES v      ON v.id = s.venue_id
              JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.PRODUCTIONS p ON p.id = s.production_id
              WHERE p.theatrical_slug=%(slug)s AND v.provider_venue_id IS NOT NULL
                AND DATE(s.local_start_time) BETWEEN %(start)s AND %(end)s
              GROUP BY 1, 2
            )
            SELECT pvid,
                   SUM(mat + prime + late) AS st,
                   COUNT(*) AS days,
                   SUM(IFF(mat > 0 AND prime > 0, 1, 0)) AS clean_days,
                   SUM(IFF(mat > 0, 1, 0)) AS mat_days,
                   SUM(mat) AS mat, SUM(prime) AS prime, SUM(late) AS late
            FROM daily GROUP BY pvid
        """, {"slug": slug, "start": start, "end": end})
        st_by_pvid = {}
        for pid, st, days, clean_days, mat_days, mat, prime, late in cur.fetchall():
            st_by_pvid[pid] = {"st": int(st), "days": int(days), "clean_days": int(clean_days),
                               "mat_days": int(mat_days), "mat": int(mat),
                               "prime": int(prime), "late": int(late)}

        # normalized venue name -> provider_venue_id, for venues that have showtimes
        # (fallback when a Mica venue doesn't resolve through ORACLEV3)
        cur.execute("""
            SELECT DISTINCT v.name, CAST(v.provider_venue_id AS STRING) AS pvid
            FROM RAW_AIRBYTE.THEATRICAL_MONOLITH.SHOWTIMES s
            JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.VENUES v      ON v.id = s.venue_id
            JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.PRODUCTIONS p ON p.id = s.production_id
            WHERE p.theatrical_slug=%(slug)s AND v.provider_venue_id IS NOT NULL
              AND DATE(s.local_start_time) BETWEEN %(start)s AND %(end)s
        """, {"slug": slug, "start": start, "end": end})
        name_to_pid = {}
        for name, pid in cur.fetchall():
            name_to_pid.setdefault(_norm(name), pid)

    def resolve_pid(m):
        return (ov_name.get(_norm(m["venue"]))
                or (ov_atom.get(str(m["atom"])) if m["atom"] else None)
                or name_to_pid.get(_norm(m["venue"])))

    ZERO = {"st": 0, "days": 0, "clean_days": 0, "mat_days": 0, "mat": 0, "prime": 0, "late": 0}

    venues = []
    for m in mica:
        pid = resolve_pid(m)
        if pid is None:
            s, flag = ZERO, "unresolved"            # venue not matched — can't confirm
        else:
            s = st_by_pvid.get(pid)
            if s is None:
                s, flag = ZERO, "no_show"           # known venue, zero showtimes
            else:
                # "Clean" = matinee + prime EVERY play-day. "Single Matinee" = a
                # matinee every play-day. Anything scheduling but short of that = under.
                pt = (m["playtype"] or "").strip().lower()
                if pt == "single matinee":
                    flag = "on_track" if s["mat_days"] == s["days"] else "under"
                else:
                    flag = "on_track" if s["clean_days"] == s["days"] else "under"
        booker = re.sub(r"\s+", " ", m["booker"]).strip() if m["booker"] else "(unassigned)"
        venues.append({
            "venue": m["venue"], "city": (m["city"] or "").strip(), "state": (m["state"] or "").strip(),
            "booker": booker,
            "buyer": (m["buyer"] or "").strip(), "buyer_email": (m["buyer_email"] or "").strip(),
            "playtype": m["playtype"] or "", "screens": m["screens"],
            "status": m["status"], "showtimes": s["st"], "days": s["days"],
            "matinee": s["mat"], "prime": s["prime"], "late": s["late"],
            "flag": flag, "pvid": pid,
        })

    # --- enrich with venue tags (provider_venue_id -> monolith UUID -> tags) ---
    pids = sorted({v["pvid"] for v in venues if v["pvid"]})
    pvid_to_uuid = {}
    if pids:
        in_list = ",".join("'" + p.replace("'", "''") + "'" for p in pids)
        with _lock:
            cur = _cursor()
            cur.execute(f"""
                SELECT CAST(provider_venue_id AS STRING) AS pvid, id AS venue_uuid
                FROM RAW_AIRBYTE.THEATRICAL_MONOLITH.VENUES
                WHERE CAST(provider_venue_id AS STRING) IN ({in_list})
                QUALIFY ROW_NUMBER() OVER (
                    PARTITION BY provider_venue_id
                    ORDER BY CASE WHEN COALESCE(status,'supported')='supported' THEN 0 ELSE 1 END,
                             updated_at DESC NULLS LAST) = 1
            """)
            pvid_to_uuid = {pid: uid for pid, uid in cur.fetchall()}
    tag_map = fetch_tags(slug, list(pvid_to_uuid.values()))
    for v in venues:
        uid = pvid_to_uuid.get(v["pvid"])
        v["tags"] = tag_map.get(uid, []) if uid else []
        if v["flag"] == "no_show" and (set(v["tags"]) & EXPLAIN_TAGS):
            v["flag"] = "expected"     # 0 showtimes explained by an online-ticketing tag

    # KPIs
    counts = {k: 0 for k in ("on_track", "under", "no_show", "expected", "unresolved")}
    for v in venues:
        counts[v["flag"]] += 1
    agreed = len(venues)
    scheduled = counts["on_track"] + counts["under"]
    kpis = {
        "agreed": agreed,
        "scheduled": scheduled,
        "scheduled_pct": round(100 * scheduled / agreed) if agreed else 0,
        "no_show": counts["no_show"],
        "under": counts["under"],
        "expected": counts["expected"],
        "unresolved": counts["unresolved"],
    }

    # by booker
    bk = {}
    for v in venues:
        b = bk.setdefault(v["booker"], {"booker": v["booker"], "agreed": 0,
                                        "scheduled": 0, "no_show": 0, "under": 0})
        b["agreed"] += 1
        if v["flag"] in ("on_track", "under"):
            b["scheduled"] += 1
        if v["flag"] == "no_show":
            b["no_show"] += 1
        if v["flag"] == "under":
            b["under"] += 1
    by_booker = sorted(bk.values(), key=lambda x: -x["no_show"])

    return {
        "title": slug, "window": {"start": start, "end": end},
        "kpis": kpis, "by_booker": by_booker, "venues": venues,
    }


def fetch_showtimes(slug, start, end, pvid):
    """Per-showtime detail for one venue (by provider_venue_id), grouped by day.

    Returns [{date_label, items:[{time, sold}]}] where sold is % occupied (or None).
    """
    with _lock:
        cur = _cursor()
        cur.execute("""
            SELECT s.local_start_time, atm.percentage_occupied
            FROM RAW_AIRBYTE.THEATRICAL_MONOLITH.SHOWTIMES s
            JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.VENUES v      ON v.id = s.venue_id
            JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.PRODUCTIONS p ON p.id = s.production_id
            LEFT JOIN RAW_AIRBYTE.THEATRICAL_MONOLITH.ATOM_SHOWTIMES atm ON atm.showtime_id = s.id
            WHERE p.theatrical_slug=%(slug)s
              AND CAST(v.provider_venue_id AS STRING)=%(pvid)s
              AND DATE(s.local_start_time) BETWEEN %(start)s AND %(end)s
            ORDER BY s.local_start_time
        """, {"slug": slug, "pvid": str(pvid), "start": start, "end": end})
        rows = cur.fetchall()

    days = []
    cur_key, cur_items = None, None
    for ts, pct in rows:
        key = ts.strftime("%A, %b %d")
        if key != cur_key:
            cur_key = key
            cur_items = []
            days.append({"date_label": key, "items": cur_items})
        sold = None
        if pct is not None:
            sold = round(pct * 100) if pct <= 1 else round(pct)
        cur_items.append({"time": ts.strftime("%I:%M %p").lstrip("0"), "sold": sold})
    return days


# ---------------------------------------------------------------------------
# Result cache + scheduled refresh
# ---------------------------------------------------------------------------

def _load_cache():
    global _cache
    if _cache is None:
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                _cache = json.load(f)
        except Exception:
            _cache = {}
    return _cache


def _save_cache():
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(_cache, f)
    except Exception as e:
        print(f"[cache] save failed: {e}", flush=True)


def get_coverage(slug, start, end, force=False):
    """Cache-aware coverage. Serves the last scheduled result instantly unless
    force=True (Refresh button), which recomputes live and updates the cache."""
    key = f"{slug}|{start}|{end}"
    if not force:
        with _cache_lock:
            cache = _load_cache()
            if key in cache:
                return cache[key]
    data = fetch_coverage(slug, start, end)
    data["generated_at"] = datetime.now(timezone.utc).isoformat()
    with _cache_lock:
        cache = _load_cache()
        cache[key] = data
        _save_cache()
    return data


_FLAG_LABEL = {"no_show": "No-show", "under": "Under", "on_track": "On track",
               "expected": "Expected", "unresolved": "Unresolved"}
_FLAG_ORDER = {"no_show": 0, "under": 1, "expected": 2, "unresolved": 3, "on_track": 4}


def _filter_venues(data, flag=None, booker=None, q=None):
    rows = data["venues"]
    if flag and flag != "all":
        rows = [v for v in rows if v["flag"] == flag]
    if booker:
        rows = [v for v in rows if v["booker"] == booker]
    if q:
        ql = q.lower()
        rows = [v for v in rows if ql in " ".join([
            v.get("venue", ""), v.get("city", ""), v.get("state", ""),
            v.get("buyer", ""), v.get("buyer_email", "")]).lower()]
    return sorted(rows, key=lambda v: (_FLAG_ORDER.get(v["flag"], 9), -(v.get("screens") or 0)))


def build_xlsx(slug, start, end, flag=None, booker=None, q=None):
    """Build an .xlsx (Summary + Venues sheets) for the given title/window/filters."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    data = get_coverage(slug, start, end)
    rows = _filter_venues(data, flag, booker, q)
    k = data["kpis"]

    wb = Workbook()
    s = wb.active
    s.title = "Summary"
    for r in [["Booking Coverage"], ["Title", slug], ["Window", f"{start} to {end}"],
              ["Generated (UTC)", data.get("generated_at", "")], [],
              ["Agreed venues", k["agreed"]],
              ["Scheduled", f'{k["scheduled"]} ({k["scheduled_pct"]}%)'],
              ["No-shows", k["no_show"]], ["Under-running", k["under"]],
              ["Expected", k["expected"]], ["Unresolved", k["unresolved"]],
              [], ["Rows in this export", len(rows)]]:
        s.append(r)
    s["A1"].font = Font(bold=True, size=14)
    s.column_dimensions["A"].width = 18
    s.column_dimensions["B"].width = 36

    ws = wb.create_sheet("Venues")
    headers = ["Venue", "City", "State", "Booker", "Buyer", "Buyer Email", "Playtype",
               "Screens", "Showtimes", "Days", "Matinee", "Prime", "Late", "Status", "Tags"]
    ws.append(headers)
    for v in rows:
        ws.append([v.get("venue", ""), v.get("city", ""), v.get("state", ""),
                   v.get("booker", ""), v.get("buyer", ""), v.get("buyer_email", ""),
                   v.get("playtype", ""), v.get("screens"), v.get("showtimes"),
                   v.get("days"), v.get("matinee", 0), v.get("prime", 0), v.get("late", 0),
                   _FLAG_LABEL.get(v["flag"], v["flag"]),
                   ", ".join(v.get("tags", []))])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate([34, 16, 8, 20, 20, 26, 12, 9, 11, 7, 8, 8, 8, 12, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def refresh_all(window_days=DEFAULT_WINDOW_DAYS):
    """Recompute + cache the default (opening-week) window for every current title.
    Called by the scheduler each morning/afternoon."""
    done = []
    for t in get_titles():
        sd = t.get("start_date")
        if not sd:
            continue
        try:
            d0 = date.fromisoformat(sd)
            start, end = d0.isoformat(), (d0 + timedelta(days=window_days)).isoformat()
            get_coverage(t["slug"], start, end, force=True)
            done.append(f"{t['slug']} {start}..{end}")
        except Exception as e:
            print(f"[refresh] {t['slug']} failed: {e}", flush=True)
    print(f"[refresh] cached {len(done)} title(s): {done}", flush=True)
    return done
